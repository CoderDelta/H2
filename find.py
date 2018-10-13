
#----------------------------------------------------#
#-                                                  -#
#-                Bogdan Polvanov                   -#
#-                 Project Euler                    -#
#-                    Task_26                       -#
#-                                                  -#
#----------------------------------------------------#

import time
import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook

start = time.time()
#-------------------- working space ---------------------


book = xlrd.open_workbook("corect.xlsx")
book1 = xlrd.open_workbook("energiy_perekhodu.xlsx")
result = []
first_sheet = book.sheet_by_index(0)

energy = []
result = []
l = 0
lm =[]
sq= []
pr = [ ] 
for n in range(1,70):# перербирає експерерментальні довжини хвиль
    stan = first_sheet.row_values(n)# відкриваємо кожний рядок book
    min_stan = float(stan[16])-float(stan[20])
    max_stan = float(stan[16]) + float(stan[20])
    sheet_name = book1.sheet_names()
    for i in range(0,len(sheet_name)):# перебирає аркуші
       # print(i)
        theor_sheet = book1.sheet_by_index(i)
        sj = theor_sheet.row_values(10)[2::]
        sj1 = theor_sheet.row_values(11)[2::]
        for fx in range(0,len(sj)):
            st = str(int(sj[fx]))+ '->' + str(int(sj1[fx]))
            energy.append(st)
        for k in range(13,theor_sheet.nrows): # проганяє експерементальну енергію через теоретичну
            check = theor_sheet.row_values(k)
            for n in range(2,len(check)):
                if min_stan <= check[n] <= max_stan:
                    nu = str(int(check[0])) + '->' + str(int(check[1])) # записуємо довжини хвиль у result
                    result.append([sheet_name[i],nu,str(energy[n-2]),round(check[n],2),round(float(stan[16]),2),stan[20]])
                    #pr.append([sheet_name[i],nu,str(energy[n-2]),round(check[n],2),round(float(stan[16]),2),stan[20]])
                    
        #corect = []

        #for ind in pr:# перевіряємо на схожість переходи
        #    corect.append(ind[3])
        #if corect == []:
        #    continue
        #elif corect != [] :
        #    for kn in  pr:
        #       if min(corect) == kn[3]:
        #           snd = []
        #           for kl in result:
        #               snd.append(kl[3])
        #           if  min(corect) not in snd:
        #               result.append(kn)
        #pr = []


wb = load_workbook('proper.xlsx')
ws1 = wb.get_sheet_by_name("Sheet1")
ws1.cell(row=1, column=1).value = 'state'
ws1.cell(row=1, column=2).value = 'nu'
ws1.cell(row=1, column=3).value = 'j'
ws1.cell(row=1, column=4).value = 'energy'
ws1.cell(row=1, column=5).value = 'eksp_energy'
ws1.cell(row=1, column=6).value = 'error'

for  m in range(2,len(result) + 2):
    ws1.cell(row=m, column=1).value = result[m-2][0]
    ws1.cell(row=m, column=2).value = result[m-2][1]
    ws1.cell(row=m, column=3).value = result[m-2][2]
    ws1.cell(row=m, column=4).value = result[m-2][3]
    ws1.cell(row=m, column=5).value = result[m-2][4]
    ws1.cell(row=m, column=6).value = result[m-2][5]
wb.save('proper.xlsx')


    





    

#-------------------    end    ---------------------
end = time.time()
print('time work - ',end - start)



